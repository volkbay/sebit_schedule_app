# TODO: Bu satirlar silinecek
# from subprocess import call
# call(["pyside2-uic", "/home/volkan/Proje/Qt/sebit/mainwindows.ui", "-o", "/home/volkan/Proje/Qt/sebit/mainwindows.py"])
# call(["pyside2-uic", "/home/volkan/Proje/Qt/sebit/personawindow.ui", "-o", "/home/volkan/Proje/Qt/sebit/personawindow.py"])
# call(["pyside2-uic", "/home/volkan/Proje/Qt/sebit/personapercentwindow.ui", "-o", "/home/volkan/Proje/Qt/sebit/personapercentwindow.py"])

from PySide2.QtWidgets import QApplication, QMainWindow
from PySide2.QtWidgets import QListWidgetItem
from PySide2.QtCore import Slot
import math
import xlrd
from openpyxl import load_workbook
import numpy as np
import sys
from personaWin import PersonaPage
from personaPerWin import PersonaPerPage
import mainwindows


class MainPage(QMainWindow, mainwindows.Ui_MainWindow):
    def __init__(self):
        super(MainPage, self).__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.openWin_persona)
        self.pushButton_3.clicked.connect(self.calculate)
        self.pushButton_4.clicked.connect(self.openWin_persona_per)
        self.radioButton.clicked.connect(self.enableEntry)
        self.radioButton_2.clicked.connect(self.disableEntry)
        self.lineEdit.editingFinished.connect(self.updateInterval)
        self.ui2 = PersonaPage()
        self.ui2.persona = [{}, {}, {}, {}]
        self.ui3 = PersonaPerPage()
        self.ui3.personaPer = [[], [], [], []]
        self.personaPerFlag = [0, 0, 0, 0]
        self.ui2.selectBtn.clicked.connect(lambda: self.updatePersonaPerFlag(0))
        self.ui2.selectBtn_2.clicked.connect(lambda: self.updatePersonaPerFlag(1))
        self.ui2.selectBtn_3.clicked.connect(lambda: self.updatePersonaPerFlag(2))
        self.ui2.selectBtn_4.clicked.connect(lambda: self.updatePersonaPerFlag(3))
        self.interval = 'hepsi'
        self.result_no = 0
        self.output = {}
        self.workbook = xlrd.open_workbook('main.xlsx')
        self.workbook2 = xlrd.open_workbook('list.xlsx')

    @Slot()
    def updateInterval(self):
        if self.lineEdit.text().isdigit():
            self.interval = int(self.lineEdit.text())
        else:
            nums = self.lineEdit.text().split(",")
            if all(x.isdigit() for x in nums):
                self.interval = [int(x) for x in nums]
            else:
                self.interval = []
                for n in nums:
                    if n.isdigit():
                        self.interval.append(int(n))
                    else:
                        digit = n.split("-")
                        if digit[0].isdigit() and digit[1].isdigit():
                            self.interval.extend(range(int(digit[0]), int(digit[1])+1))
                        else:
                            self.interval = 'error'

    @Slot()
    def enableEntry(self):
        self.lineEdit.setEnabled(True)
        self.radioButton_2.setChecked(False)

    @Slot()
    def disableEntry(self):
        self.lineEdit.setEnabled(False)
        self.radioButton.setChecked(False)
        self.interval = 'hepsi'

    @Slot()
    def openWin_persona(self):
        self.ui2.exec_()
        self.updatePersonaPer()

    @Slot()
    def openWin_persona_per(self):
        self.ui3.updatePerPerWin(self.ui2.persona)
        self.ui3.exec_()

    def initTeach(self):
        self.teach = {}
        self.worksheet = self.workbook.sheet_by_name('Öğretmen Listesi')
        for row in range(0, self.worksheet.nrows):
            self.teach[self.worksheet.cell(row, 0).value] = 0

    def initCourse(self):
        self.opt_course_list = {}
        self.worksheet = self.workbook.sheet_by_name('Seçmeli')
        for row in range(0, self.worksheet.nrows):
            self.opt_course_list[self.worksheet.cell(row, 0).value] = [self.worksheet.cell(row, 1).value, int(self.worksheet.cell(row, 2).value), 0]

    def emptyCourse(self):
        for k in self.opt_course_list.keys():
            self.opt_course_list[k][-1] = 0

    @Slot()
    def updatePersonaPerFlag(self, tab):
        if not self.ui2.persona[tab]:
            self.personaPerFlag[tab] = 1
        else:
            self.personaPerFlag[tab] = 2

    def updatePersonaPer(self):
        for i in range(4):
            num = len(self.ui2.persona[i])
            if self.personaPerFlag[i] == 1 and num != 0:
                self.personaPerFlag[i] = 3
                self.ui3.personaPer[i] = [math.floor(100/num)] * num
                if num != 1:
                    self.ui3.personaPer[i][-1] = 100 - sum(self.ui3.personaPer[i][0:-1])
            elif self.personaPerFlag[i] == 2:
                self.ui3.personaPer[i] = self.ui3.personaPer[i] + ([0] * (num-len(self.ui3.personaPer[i])))

    @Slot()
    def calculate(self):
        self.listWidget.clear()
        self.listWidget.addItem("Hesaplamaya başlıyor. Okul listesi - " + str(self.interval))
        self.initCourse()
        self.worksheet2 = self.workbook2.sheet_by_index(0)
        numSchool = self.worksheet2.nrows - 3
        if self.interval == 'error':
            self.listWidget.clear()
            self.listWidget.addItem("Doğru bir aralık giriniz.")
        elif self.interval == 'hepsi':
            self.init_result()
            for school in range(1, numSchool+1):
                self.calculate_loop(school)
        elif isinstance(self.interval, int):
            self.init_result()
            if self.interval < numSchool + 1:
                self.calculate_loop(self.interval)
            else:
                self.listWidget.clear()
                self.listWidget.addItem("Okul sayısını aştınız.")
        else:
            self.init_result()
            for school in self.interval:
                if school < numSchool + 1:
                    self.calculate_loop(school)
        self.workbook3.save('result.xlsx')

    def calculate_loop(self, school_ind):
        school_info = self.worksheet2.row_values(2+school_ind, 0, 31)
        self.initTeach()
        self.emptyCourse()
        self.num_of_rooms = 0
        self.numStudent = [int(school_info[9]), int(school_info[11]), int(school_info[13]), int(school_info[15])]
        self.listWidget.addItem("#Okul Sıra No : " + str(school_ind))
        self.listWidget.addItem("#Öğrenci Mevcuları : " + str(self.numStudent))
        self.perTeach = int(self.lineEdit_dersyuku.text())
        self.capacity = int(self.lineEdit_derslikkap.text())
        self.perPersona = [[], [], [], []]
        calculate_flag = True
        for tab in range(4):
            persona_array = np.array(self.ui3.personaPer[tab])
            if persona_array.size == 0:
                self.perPersona[tab] = np.array([0])
                calculate_flag = False
            else:
                self.perPersona[tab] = (self.numStudent[tab] * persona_array / 100).round()
                if persona_array.size != 1:
                    self.perPersona[tab][-1] = self.numStudent[tab] - self.perPersona[tab][0:-1].sum()
                self.calMust(tab)
                self.calElec_init(tab)
                calculate_flag = True
        if calculate_flag:
            self.calcElec_start()
        self.num_of_rooms = math.ceil(sum(self.teach.values()) / 40)
        self.listWidget.addItem("#Sınıf Sayısı (asgari): " + str(self.num_of_rooms))
        self.listWidget.addItem("\n---------")
        self.output = {k: math.ceil(v/self.perTeach) for k, v in self.teach.items()}
        for k, v in self.output.items():
            self.listWidget.addItem(QListWidgetItem(k+':'+str(v)))
        self.listWidget.addItem("\n---------")
        self.write_result(school_info)

    def freeTeach(self, lst):
        lst2 = []  # Son eklenen hocanın haftalik yuku
        lst3 = []  # Bir branstaki toplam yuk
        lst5 = []  # Son hocada en az yük olan derslerin toplam brans yukleri
        for word in lst:
            lst2.append(self.teach[word] % self.perTeach)
            lst3.append(self.teach[word])
        lst4 = np.where(lst2 == np.min(lst2))  # Indexes of all minimum elements
        lst4 = lst4[0]  # Yapi sebebiyle dummy line
        if lst4.size == 1:
            self.free = lst[lst4[0]]
        else:
            lst5 = np.array([lst3[i] for i in lst4])
            self.free = lst[lst4[np.argmin(lst5)]]

    def calMust(self, tab):
        self.worksheet = self.workbook.sheet_by_index(tab + 1)
        self.numCourse = self.worksheet.nrows
        persona_ind = 0
        for v in self.ui2.persona[tab].values():
            numClass = math.ceil(self.perPersona[tab][persona_ind]/self.capacity)
            persona_ind = persona_ind + 1
            for row in range(1, self.numCourse):
                if v[row-1] == 2:
                    for i in range(0, numClass):
                        word = self.worksheet.cell(row, 2).value
                        word_list = word.split(",")
                        if len(word_list) > 1:
                            self.freeTeach(word_list)
                            self.teach[self.free] = self.teach[self.free] + int(self.worksheet.cell(row, 3).value)
                        elif word == "Hepsi":
                            self.freeTeach(list(self.teach.keys()))
                            self.teach[self.free] = self.teach[self.free] + int(self.worksheet.cell(row, 3).value)
                        else:
                            self.teach[word] = self.teach[word] + int(self.worksheet.cell(row, 3).value)

    def calElec_init(self, tab):
        self.worksheet = self.workbook.sheet_by_index(tab + 1)
        persona_ind = 0
        for v in self.ui2.persona[tab].values():
            for row in range(1, self.numCourse):
                if v[row-1] == 1:
                    name = self.worksheet.cell(row, 0).value
                    self.opt_course_list[name][-1] = self.opt_course_list[name][-1] + self.perPersona[tab][persona_ind]
        persona_ind = persona_ind + 1

    def calcElec_start(self):
        for course in self.opt_course_list.keys():
            word = self.opt_course_list[course][0]
            word_list = word.split(",")
            numClass = math.ceil(self.opt_course_list[course][2]/self.capacity)
            for i in range(0, numClass):
                if len(word_list) > 1:
                    self.freeTeach(word_list)
                    self.teach[self.free] = self.teach[self.free] + self.opt_course_list[course][1]
                elif word == "Hepsi":
                    self.freeTeach(list(self.teach.keys()))
                    self.teach[self.free] = self.teach[self.free] + self.opt_course_list[course][1]
                else:
                    self.teach[word] = self.teach[word] + self.opt_course_list[course][1]

    def init_result(self):
        self.workbook3 = load_workbook(filename='result.xlsx')
        self.result_no = len(self.workbook3.sheetnames)
        sheetname = "Result_" + str(self.result_no)
        self.workbook3.create_sheet(sheetname)
        sheet = self.workbook3[sheetname]
        sheet.cell(row=1, column=1, value="PERSONA ve YÜZDELERİ")
        r = 0
        for tab in self.ui2.persona:
            c = 0
            for n in tab.keys():
                sheet.cell(row=r + 2, column=c + 1, value=n+" : "+str(self.ui3.personaPer[r][c]))
                c = c + 1
            r = r + 1
        sheet.cell(row=r + 3, column=1, value="OKUL LİSTESİ")
        sheet.cell(row=r + 4, column=1, value=str(self.interval))
        sheet.cell(row=r + 6, column=1, value="SONUÇLAR")
        self.titles = ["Okul Sıra No", "9.Sınıf Mevcut", "10.Sınıf Mevcut", "11.Sınıf Mevcut", "12.Sınıf Mevcut",
                       "Mevcut Sınıf Sayısı", "Asgari Sınıf Sayısı", "Toplam Öğretmen Sayısı", "Asgari Öğretmen Sayısı"]
        c = 0
        for title in self.titles:
            sheet.cell(row=r + 7, column=c + 1, value=title)
            c = c + 1
        self.initTeach()
        for title in self.teach.keys():
            sheet.cell(row=r + 7, column=c + 1, value=title)
            c = c + 1

    def write_result(self, school_info):
        sheetname = "Result_" + str(self.result_no)
        sheet = self.workbook3[sheetname]
        results = [int(school_info[0]), int(school_info[9]), int(school_info[11]), int(school_info[13]), int(school_info[15]),
                   int(school_info[5]), self.num_of_rooms, int(school_info[4]), sum(self.output.values())]
        r = sheet.max_row
        c = 0
        for res in results:
            sheet.cell(row=r + 1, column=c + 1, value=res)
            c = c + 1
        for n, res in enumerate(self.output.values()):
            if school_info[n + 17] == '':
                school_info[n + 17] = 0
            sheet.cell(row=r + 1, column=c + 1, value=str(int(school_info[n + 17])) + " (" + str(int(res)) + ")")
            c = c + 1
        sheet.cell(row=r + 1, column=c, value=str(0) + " (" + str(int(res)) + ")")  # PDR verisi olmadığı için düzeltme


app = QApplication(sys.argv)
ui = MainPage()
ui.show()
sys.exit(app.exec_())
