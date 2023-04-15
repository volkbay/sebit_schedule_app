from PySide2.QtWidgets import (QDialog, QLineEdit, QWidget, QTableWidget, QPushButton,
                               QSizePolicy, QVBoxLayout, QTableWidgetItem, QScrollArea)
from PySide2.QtCore import (QSize, Slot)
import numpy as np
import xlrd
from openpyxl import load_workbook
import personawindow


class PersonaPage(QDialog, personawindow.Ui_PersonaWindow):
    def __init__(self):
        super(PersonaPage, self).__init__()
        self.setupUi(self)
        self.updateDropMenu()
        self.slot_no = np.zeros(4, dtype=int)
        num_p = 20
        self.nameLine = np.empty((4, num_p), dtype=QLineEdit)
        self.group_box = np.empty((4, num_p), dtype=QWidget)
        self.table = np.empty((4, num_p), dtype=QTableWidget)
        self.saveBtn = np.empty((4, num_p), dtype=QPushButton)
        self.selectBtn.clicked.connect(lambda: self.initPersona(self.dropLst.currentIndex(), 0))
        self.selectBtn_2.clicked.connect(lambda: self.initPersona(self.dropLst_2.currentIndex(), 1))
        self.selectBtn_3.clicked.connect(lambda: self.initPersona(self.dropLst_3.currentIndex(), 2))
        self.selectBtn_4.clicked.connect(lambda: self.initPersona(self.dropLst_4.currentIndex(), 3))
        self.confirmBtn.clicked.connect(self.loadPersona)

    def updateDropMenu(self):
        book = load_workbook(filename='main.xlsx')
        sheet = book["9.Sınıf"]
        self.dropLst.clear()
        for c in range(5, sheet.max_column + 1):
            self.dropLst.addItem(sheet.cell(row=1, column=c).value)
        sheet = book["10.Sınıf"]
        self.dropLst_2.clear()
        for c in range(5, sheet.max_column + 1):
            self.dropLst_2.addItem(sheet.cell(row=1, column=c).value)
        sheet = book["11.Sınıf"]
        self.dropLst_3.clear()
        for c in range(5, sheet.max_column + 1):
            self.dropLst_3.addItem(sheet.cell(row=1, column=c).value)
        sheet = book["12.Sınıf"]
        self.dropLst_4.clear()
        for c in range(5, sheet.max_column + 1):
            self.dropLst_4.addItem(sheet.cell(row=1, column=c).value)

    @Slot()
    def initPersona(self, index, tab):
        self.workbook = xlrd.open_workbook('main.xlsx')
        self.worksheet = self.workbook.sheet_by_index(tab + 1)
        self.group_box[tab, self.slot_no[tab]] = QWidget(self.gradeTab.widget(tab).findChild(QScrollArea).widget())
        self.group_box[tab, self.slot_no[tab]].setSizePolicy(QSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed))
        self.group_box[tab, self.slot_no[tab]].setMinimumSize(QSize(500, 400))
        self.gradeTab.widget(tab).findChild(QScrollArea).widget().layout().insertWidget(self.slot_no[tab], self.group_box[tab, self.slot_no[tab]])
        self.nameLine[tab, self.slot_no[tab]] = QLineEdit(self.group_box[tab, self.slot_no[tab]])
        self.nameLine[tab, self.slot_no[tab]].setText(self.worksheet.cell(0, 4+index).value)
        self.table[tab, self.slot_no[tab]] = QTableWidget(self.group_box[tab, self.slot_no[tab]])
        self.table[tab, self.slot_no[tab]].setColumnCount(3)
        self.table[tab, self.slot_no[tab]].setRowCount(self.worksheet.nrows-1)
        self.table[tab, self.slot_no[tab]].horizontalHeader().setVisible(False)
        self.fillTable(index, tab)
        self.table[tab, self.slot_no[tab]].resizeColumnsToContents()
        self.saveBtn[tab, self.slot_no[tab]] = QPushButton(self.group_box[tab, self.slot_no[tab]])
        self.saveBtn[tab, self.slot_no[tab]].setText("Kaydet")
        dummy = self.slot_no[tab]
        self.saveBtn[tab, self.slot_no[tab]].clicked.connect(lambda: self.savePersona(dummy, tab))
        group_layout = QVBoxLayout(self.group_box[tab, self.slot_no[tab]])
        group_layout.addWidget(self.table[tab, self.slot_no[tab]])
        group_layout.addWidget(self.nameLine[tab, self.slot_no[tab]])
        group_layout.addWidget(self.saveBtn[tab, self.slot_no[tab]])
        self.slot_no[tab] = self.slot_no[tab] + 1

    @Slot()
    def savePersona(self, btn_no, tab):
        if self.nameLine[tab, btn_no].isEnabled():
            self.nameLine[tab, btn_no].setEnabled(False)
            self.table[tab, btn_no].setEnabled(False)
            self.saveBtn[tab, btn_no].setText("Düzenle")
            book = load_workbook(filename='main.xlsx')
            sheetnames = book.sheetnames
            sheet = book[sheetnames[tab + 1]]
            new_persona = True
            for c in range(5, sheet.max_column + 1):
                if sheet.cell(row=1, column=c).value == self.nameLine[tab, btn_no].text():
                    new_persona = False
                    for r in range(2, sheet.max_row + 1):
                        sheet.cell(row=r, column=c, value=self.table[tab, btn_no].item(r-2, 1).text())
            if new_persona:
                new_col = sheet.max_column + 1
                sheet.cell(row=1, column=new_col, value=self.nameLine[tab, btn_no].text())
                for r in range(2, sheet.max_row + 1):
                    sheet.cell(row=r, column=new_col, value=self.table[tab, btn_no].item(r-2, 1).text())
            book.save('main.xlsx')
            self.updateDropMenu()

        else:
            self.nameLine[tab, btn_no].setEnabled(True)
            self.table[tab, btn_no].setEnabled(True)
            self.saveBtn[tab, btn_no].setText("Kaydet")

    def fillTable(self, index, tab):
        self.workbook = xlrd.open_workbook('main.xlsx')
        self.worksheet = self.workbook.sheet_by_index(tab + 1)
        for row in range(1, self.worksheet.nrows):
            item = QTableWidgetItem(self.worksheet.cell(row, 0).value)
            self.table[tab, self.slot_no[tab]].setItem(row-1, 0, item)
            item = QTableWidgetItem(self.worksheet.cell(row, index+4).value)
            self.table[tab, self.slot_no[tab]].setItem(row-1, 1, item)
            item = QTableWidgetItem(str(self.worksheet.cell(row, 3).value))
            self.table[tab, self.slot_no[tab]].setItem(row-1, 2, item)

    @Slot()
    def loadPersona(self):
        self.workbook = xlrd.open_workbook('main.xlsx')
        self.persona = [{}, {}, {}, {}]
        for tab in range(4):
            self.worksheet = self.workbook.sheet_by_index(tab + 1)
            for n in range(0, self.slot_no[tab]):
                self.persona[tab][self.nameLine[tab, n].text()] = np.empty(self.worksheet.nrows-1, dtype=int)
                for row in range(1, self.worksheet.nrows):
                    if self.table[tab, n].item(row-1, 1).text() in ('Z', 'z'):
                        self.persona[tab][self.nameLine[tab, n].text()][row-1] = 2
                    elif self.table[tab, n].item(row-1, 1).text() in ('S', 's'):
                        self.persona[tab][self.nameLine[tab, n].text()][row-1] = 1
                    else:
                        self.persona[tab][self.nameLine[tab, n].text()][row-1] = 0
        self.hide()
